import tkinter as tk
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import xlwt
import time

app = tk.Tk()
app.title("Student Result Automation")

# Tkinter input variables
driver_path = tk.StringVar()
usn_file = tk.StringVar()
website = tk.StringVar()
save_path = tk.StringVar()
start_row = tk.StringVar()
end_row = tk.StringVar()


def pick_file(target, file_types) -> None:
    """Open a file picker and store the selected path."""
    path = filedialog.askopenfilename(title="Select a File", filetypes=file_types)
    if path:
        target.set(path)


def choose_driver() -> None:
    pick_file(driver_path, [("Executables", "*.exe"), ("All Files", "*.*")])


def choose_usn_file() -> None:
    pick_file(usn_file, [("Excel Files", "*.xlsx"), ("All Files", "*.*")])


def choose_save_path() -> None:
    # Keeping original behavior (selecting an existing file path)
    pick_file(save_path, [("Excel Files", "*.xls"), ("All Files", "*.*")])


def start() -> None:
    dpath = driver_path.get().strip()
    input_file = usn_file.get().strip()
    url = website.get().strip()
    output_file = save_path.get().strip()

    try:
        start_i = int(start_row.get())
        end_i = int(end_row.get())
    except ValueError:
        print("Start/End row must be numbers.")
        return

    # Output workbook
    out_book = xlwt.Workbook()
    out_sheet = out_book.add_sheet("Sheet1", cell_overwrite_ok=True)
    out_book.add_sheet("sheet2", cell_overwrite_ok=True)

    orange = xlwt.easyxf("pattern: pattern solid, fore_colour orange")

    # Input workbook
    try:
        in_book = openpyxl.load_workbook(input_file)
        in_sheet = in_book.active
    except Exception as e:
        print(f"Failed to open input Excel file: {e}")
        return

    time.sleep(3)

    # Selenium setup
    service = Service(dpath)
    opts = Options()
    opts.add_experimental_option("detach", True)

    driver = webdriver.Chrome(service=service, options=opts)
    actions = ActionChains(driver)
    driver.maximize_window()
    driver.get(url)

    main_win = driver.window_handles[0]
    time.sleep(3)

    # Main page elements
    try:
        usn_box = driver.find_element(By.XPATH, '//*[@id="raj"]/div[1]/div/input')
        captcha_box = driver.find_element(By.XPATH, '//*[@id="raj"]/div[2]/div[1]/input')
        submit_btn = driver.find_element(By.XPATH, '//*[@id="submit"]')
    except NoSuchElementException:
        print("Page elements not found. The website layout may have changed.")
        return

    time.sleep(3)
    captcha = input("Enter captcha shown in the browser: ")

    print(f"Processing rows {start_i} to {end_i}...")

    for row in range(start_i, end_i + 1):
        cell = in_sheet.cell(row=row, column=1)
        usn = cell.value

        if not usn:
            print(f"Skipping row {row}: empty USN")
            continue

        usn_box.send_keys(usn)
        time.sleep(0.5)
        captcha_box.send_keys(captcha)
        time.sleep(0.5)

        # Open result in new tab/window
        actions.key_down(Keys.LEFT_CONTROL).perform()
        submit_btn.click()
        actions.key_up(Keys.LEFT_CONTROL).perform()

        if len(driver.window_handles) > 1:
            result_win = driver.window_handles[1]
            driver.switch_to.window(result_win)
        else:
            print("Result window did not open.")
            usn_box.clear()
            continue

        try:
            page_usn = driver.find_element(
                By.XPATH,
                '//*[@id="dataPrint"]/div[2]/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr[1]/td[2]'
            ).text
            name = driver.find_element(
                By.XPATH,
                '//*[@id="dataPrint"]/div[2]/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr[2]/td[2]'
            ).text

            subjects = []
            for k in range(1, 15):
                try:
                    base = (
                        '//*[@id="dataPrint"]/div[2]/div/div/div[2]/div[1]/div/div/div[2]/div/div/div[2]/div'
                        f'/div[{k+1}]'
                    )
                    sub_name = driver.find_element(By.XPATH, f"{base}/div[1]").text
                    ia = driver.find_element(By.XPATH, f"{base}/div[3]").text
                    see = driver.find_element(By.XPATH, f"{base}/div[4]").text
                    total = driver.find_element(By.XPATH, f"{base}/div[5]").text
                    res = driver.find_element(By.XPATH, f"{base}/div[6]").text

                    subjects.append(
                        {"name": sub_name, "ia": ia, "see": see, "total": total, "res": res}
                    )
                except NoSuchElementException:
                    break

            # Headers
            out_sheet.write(0, 0, "USN")
            out_sheet.write(0, 2, "NAME")

            # Student info (keeps original indexing behavior)
            out_sheet.write(row + 1, 0, page_usn)
            out_sheet.write(row + 1, 2, name)

            col = 4
            for sub in subjects:
                out_sheet.write(0, col, sub["name"])
                out_sheet.write(1, col, "IA")
                out_sheet.write(1, col + 1, "SEE")
                out_sheet.write(1, col + 2, "TOTAL")
                out_sheet.write(1, col + 3, "RES")

                out_sheet.write(row + 1, col, int(sub["ia"]))
                out_sheet.write(row + 1, col + 1, int(sub["see"]))
                out_sheet.write(row + 1, col + 2, int(sub["total"]))
                out_sheet.write(row + 1, col + 3, sub["res"], orange)

                col += 4

        except Exception as e:
            print(f"Error processing {usn}: {e}")

        driver.close()
        driver.switch_to.window(main_win)

        usn_box.clear()
        time.sleep(0.5)

    out_book.save(output_file)
    print(f"Saved results to {output_file}")


# GUI layout
tk.Label(app, text="CHROMEDRIVER PATH:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=10)
tk.Entry(app, textvariable=driver_path, width=40).grid(row=0, column=1, padx=10)
tk.Button(app, text="CHOOSE", command=choose_driver).grid(row=0, column=2, padx=10)

tk.Label(app, text="USN FILE PATH:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=10)
tk.Entry(app, textvariable=usn_file, width=40).grid(row=1, column=1, padx=10)
tk.Button(app, text="CHOOSE", command=choose_usn_file).grid(row=1, column=2, padx=10)

tk.Label(app, text="WEBSITE ADDRESS:").grid(row=3, column=0, sticky=tk.W, padx=10, pady=10)
tk.Entry(app, textvariable=website, width=40).grid(row=3, column=1, padx=10)

tk.Label(app, text="SAVE PATH:").grid(row=4, column=0, sticky=tk.W, padx=10, pady=10)
tk.Entry(app, textvariable=save_path, width=40).grid(row=4, column=1, padx=10)
tk.Button(app, text="CHOOSE", command=choose_save_path).grid(row=4, column=2, padx=10)

tk.Label(app, text="USN START (Row):").grid(row=5, column=0, sticky=tk.W, padx=10, pady=10)
tk.Entry(app, textvariable=start_row, width=20).grid(row=5, column=1, padx=10, sticky=tk.W)

tk.Label(app, text="USN END (Row):").grid(row=6, column=0, sticky=tk.W, padx=10, pady=10)
tk.Entry(app, textvariable=end_row, width=20).grid(row=6, column=1, padx=10, sticky=tk.W)

tk.Button(app, text="SUBMIT", command=start, width=15, bg="#dddddd").grid(row=7, column=0, pady=20, padx=10)
tk.Button(app, text="QUIT", command=app.quit, width=15, bg="#ffcccc").grid(row=7, column=1, pady=20, padx=10)

tk.Label(app, text="Original by: Samarth Kashyap\nDepartment of CSE").grid(row=8, column=2, pady=10)

app.mainloop()
