"""
Excel I/O module for VTU Result Automation
Handles reading USN values from input file and writing results to output file
"""

import openpyxl
import xlwt
from config import (
    EXCEL_HEADER_ROW,
    EXCEL_SUBHEADER_ROW,
    EXCEL_USN_COLUMN,
    EXCEL_NAME_COLUMN,
    EXCEL_SUBJECTS_START_COLUMN
)


def load_input_workbook(file_path):
    """
    Load the input Excel workbook containing USN values
    
    Args:
        file_path: Path to the input Excel file
        
    Returns:
        Tuple of (workbook, active_sheet) or (None, None) on error
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return workbook, sheet
    except Exception as e:
        print(f"Failed to open input Excel file: {e}")
        return None, None


def read_usn(sheet, row_index):
    """
    Read USN value from a specific row in the input sheet
    
    Args:
        sheet: The active worksheet
        row_index: Row number (1-indexed)
        
    Returns:
        USN string or None if empty
    """
    cell = sheet.cell(row=row_index, column=1)
    return cell.value


def create_output_workbook():
    """
    Create a new output workbook for storing results
    
    Returns:
        Tuple of (workbook, sheet, orange_style)
    """
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Sheet1", cell_overwrite_ok=True)
    workbook.add_sheet("sheet2", cell_overwrite_ok=True)
    
    # Orange style for result column highlighting
    orange_style = xlwt.easyxf("pattern: pattern solid, fore_colour orange")
    
    return workbook, sheet, orange_style


def write_headers(sheet):
    """
    Write static headers to the output Excel sheet
    
    Args:
        sheet: The output worksheet
    """
    sheet.write(EXCEL_HEADER_ROW, EXCEL_USN_COLUMN, "USN")
    sheet.write(EXCEL_HEADER_ROW, EXCEL_NAME_COLUMN, "NAME")


def write_student_info(sheet, row_index, usn, name):
    """
    Write student USN and name to the output sheet
    
    Args:
        sheet: The output worksheet
        row_index: Row number to write to
        usn: Student USN
        name: Student name
    """
    sheet.write(row_index + 1, EXCEL_USN_COLUMN, usn)
    sheet.write(row_index + 1, EXCEL_NAME_COLUMN, name)


def write_subject_data(sheet, row_index, subjects, orange_style):
    """
    Write subject data (headers and marks) to the output sheet
    
    Args:
        sheet: The output worksheet
        row_index: Row number to write student data to
        subjects: List of subject dictionaries containing name, ia, see, total, res
        orange_style: Excel style for highlighting result cells
    """
    col = EXCEL_SUBJECTS_START_COLUMN
    
    for sub in subjects:
        # Write subject headers
        sheet.write(EXCEL_HEADER_ROW, col, sub["name"])
        sheet.write(EXCEL_SUBHEADER_ROW, col, "IA")
        sheet.write(EXCEL_SUBHEADER_ROW, col + 1, "SEE")
        sheet.write(EXCEL_SUBHEADER_ROW, col + 2, "TOTAL")
        sheet.write(EXCEL_SUBHEADER_ROW, col + 3, "RES")
        
        # Write student marks for this subject
        sheet.write(row_index + 1, col, int(sub["ia"]))
        sheet.write(row_index + 1, col + 1, int(sub["see"]))
        sheet.write(row_index + 1, col + 2, int(sub["total"]))
        sheet.write(row_index + 1, col + 3, sub["res"], orange_style)
        
        col += 4


def save_workbook(workbook, file_path):
    """
    Save the output workbook to a file
    
    Args:
        workbook: The workbook to save
        file_path: Path where the file should be saved
    """
    workbook.save(file_path)
    print(f"Saved results to {file_path}")
